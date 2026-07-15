import base64
import csv
import io
import mimetypes
import os
import re

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from pypdf.errors import PdfReadError

FILE_ICONS = {
    "pdf": "📄",
    "image": "🖼️",
    "word": "📝",
    "excel": "📊",
    "text": "📃",
    "other": "📎",
}

MAX_UPLOAD_SIZE = 15 * 1024 * 1024
MAX_CHAT_PREVIEW_CHARS = 280_000
MAX_EXTRACTED_TEXT_CHARS = 120_000

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".jfif", ".jpe", ".heic", ".heif")
DOCUMENT_EXTENSIONS = (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".md", ".log", ".rtf")


def safe_filename(filename):
    name = os.path.basename((filename or "ek").strip()) or "ek"
    name = name.replace("\x00", "")
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    return name[:180] or "ek"


def guess_mimetype(filename, mimetype=""):
    cleaned = (mimetype or "").split(";")[0].strip().lower()
    if cleaned and cleaned != "application/octet-stream":
        return cleaned

    guessed, _ = mimetypes.guess_type(filename or "")
    return guessed or "application/octet-stream"


def get_file_category(filename, mimetype=""):
    ext = os.path.splitext(filename or "")[1].lower()
    mime = guess_mimetype(filename, mimetype)

    if ext in IMAGE_EXTENSIONS or mime.startswith("image/"):
        return "image"
    if ext == ".pdf" or mime == "application/pdf":
        return "pdf"
    if ext in (".doc", ".docx") or "word" in mime:
        return "word"
    if ext in (".xls", ".xlsx", ".csv") or "spreadsheet" in mime or "excel" in mime:
        return "excel"
    if ext in (".txt", ".md", ".log", ".rtf") or mime.startswith("text/"):
        return "text"
    return "other"


def file_icon_for(category):
    return FILE_ICONS.get(category, FILE_ICONS["other"])


def _read_file_bytes(file_storage):
    if hasattr(file_storage, "seek"):
        file_storage.seek(0)
    data = file_storage.read()
    if hasattr(file_storage, "seek"):
        file_storage.seek(0)
    return data


def pdf_bytes_to_text(data):
    if not data:
        return ""

    try:
        reader = PdfReader(io.BytesIO(data), strict=False)
    except PdfReadError as exc:
        raise ValueError("PDF okunamadı. Dosya şifreli veya bozuk olabilir.") from exc
    except Exception as exc:
        raise ValueError(f"PDF okunamadı: {exc}") from exc

    if getattr(reader, "is_encrypted", False):
        try:
            if reader.decrypt("") == 0:
                raise ValueError("PDF şifre korumalı. Şifresiz bir kopya gönderin.")
        except Exception as exc:
            raise ValueError("PDF şifre korumalı. Şifresiz bir kopya gönderin.") from exc

    parts = []
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if page_text.strip():
            parts.append(page_text)

    text = "\n".join(parts).strip()
    if not text:
        raise ValueError(
            "PDF'den metin çıkarılamadı. Tarama (görsel) PDF ise ekran görüntüsü olarak yükleyin."
        )
    return text[:MAX_EXTRACTED_TEXT_CHARS]


def pdf_to_text(file_storage):
    return pdf_bytes_to_text(_read_file_bytes(file_storage))


def docx_to_text(file_storage):
    try:
        data = _read_file_bytes(file_storage)
        doc = Document(io.BytesIO(data))
    except Exception as exc:
        raise ValueError(f"Word dosyası okunamadı: {exc}") from exc

    parts = [paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()]
    text = "\n".join(parts).strip()
    if not text:
        raise ValueError("Word dosyasından metin çıkarılamadı.")
    return text[:MAX_EXTRACTED_TEXT_CHARS]


def excel_to_text(file_storage, filename=""):
    ext = os.path.splitext(filename)[1].lower()
    data = _read_file_bytes(file_storage)

    try:
        if ext == ".csv":
            text = data.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(text))
            return "\n".join(", ".join(row) for row in reader).strip()[:MAX_EXTRACTED_TEXT_CHARS]

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in workbook.worksheets:
            parts.append(f"[{sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    parts.append(" | ".join(cells))
        workbook.close()
    except Exception as exc:
        raise ValueError(f"Excel/CSV dosyası okunamadı: {exc}") from exc

    text = "\n".join(parts).strip()
    if not text:
        raise ValueError("Excel/CSV dosyasından veri çıkarılamadı.")
    return text[:MAX_EXTRACTED_TEXT_CHARS]


def text_file_to_text(file_storage):
    text = _read_file_bytes(file_storage).decode("utf-8", errors="ignore").strip()
    if not text:
        raise ValueError("Metin dosyası boş.")
    return text[:MAX_EXTRACTED_TEXT_CHARS]


def image_to_data_url(file_storage, include_preview=True):
    filename = getattr(file_storage, "filename", "") or ""
    mime_type = guess_mimetype(filename, getattr(file_storage, "mimetype", ""))
    if not mime_type.startswith("image/"):
        mime_type = "image/jpeg"

    raw = _read_file_bytes(file_storage)
    if not raw:
        raise ValueError("Görsel dosyası boş.")

    if len(raw) > MAX_UPLOAD_SIZE:
        raise ValueError("Görsel dosyası 15 MB sınırını aşıyor.")

    encoded = base64.b64encode(raw).decode("ascii")
    data_url = f"data:{mime_type};base64,{encoded}"
    if include_preview and len(data_url) > MAX_CHAT_PREVIEW_CHARS:
        return None
    return data_url


def extract_file_text(file_storage, filename, mimetype=""):
    category = get_file_category(filename, mimetype)
    if category == "pdf":
        return pdf_to_text(file_storage), category
    if category == "word":
        return docx_to_text(file_storage), category
    if category == "excel":
        return excel_to_text(file_storage, filename), category
    if category == "text":
        return text_file_to_text(file_storage), category
    if category == "image":
        return None, category
    raise ValueError(
        f"Desteklenmeyen dosya türü: {safe_filename(filename)}. "
        "PDF, JPG, PNG, Word, Excel, CSV veya TXT gönderin."
    )


def build_file_meta(filename, category, preview=None):
    meta = {
        "name": safe_filename(filename),
        "type": category,
        "icon": file_icon_for(category),
    }
    if preview and len(preview) <= MAX_CHAT_PREVIEW_CHARS:
        meta["preview"] = preview
    return meta


def parse_uploaded_attachment(uploaded_file, max_size=MAX_UPLOAD_SIZE):
    if not uploaded_file or not uploaded_file.filename:
        return None

    filename = safe_filename(uploaded_file.filename)
    data = _read_file_bytes(uploaded_file)
    if not data:
        raise ValueError("Boş dosya gönderilemez.")

    if len(data) > max_size:
        raise ValueError(f"Dosya boyutu {max_size // (1024 * 1024)} MB sınırını aşıyor.")

    return {
        "filename": filename,
        "mimetype": guess_mimetype(filename, uploaded_file.mimetype),
        "data": data,
    }


def validate_chat_upload(uploaded_file):
    if not uploaded_file or not uploaded_file.filename:
        raise ValueError("Dosya seçilmedi.")

    filename = safe_filename(uploaded_file.filename)
    category = get_file_category(filename, uploaded_file.mimetype)
    if category == "other":
        raise ValueError(
            f"Desteklenmeyen dosya türü: {filename}. "
            "PDF, JPG, PNG, Word, Excel, CSV veya TXT gönderin."
        )

    data = _read_file_bytes(uploaded_file)
    if not data:
        raise ValueError("Boş dosya yüklenemez.")
    if len(data) > MAX_UPLOAD_SIZE:
        raise ValueError("Dosya boyutu 15 MB sınırını aşıyor.")

    return filename, category, data
